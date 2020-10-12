# Create a summary report of produce and total sold as shown below:


import openpyxl as xl
import pandas as pd

wb = xl.load_workbook('produceSales.xlsx')


ws = wb.active

maxC = ws.max_column
maxR = ws.max_row

produce_dict = {}

##print("Produce","\t","Cost Per Pound","\t","Amt Sold","\t","Total")
    
for currentrow in ws.iter_rows(min_row = 2, max_row=maxR , max_col=maxC):
    name = currentrow[0].value
    cost = float(currentrow[1].value)
    amt_sold = round(float(currentrow[2].value),2)
    total = round(float(currentrow[3].value),2)

    produce_dict.setdefault(name,[0.00,0.00,0.00])

    produce_dict[name][0] = round(cost,2)

    
    produce_dict[name][1] += round(amt_sold,2)

    produce_dict[name][2] += round(total,2)



    #print(produce_dict)
    #input()

for k,v in produce_dict.items():
    produce_dict[k] = [v[0],round(v[1]),round(v[2])]



produce_df = pd.DataFrame(produce_dict, index=["Cost Per Pound", "Quantity Sold","Total Sale"])

## display produce with smallest and largest total sales
print("Smallest Total Sales:", (produce_df.T)["Total Sale"].idxmin(), " - ", "${:,.2f}".format((produce_df.T)["Total Sale"].min()))
print("Largest Total Sales:",(produce_df.T)["Total Sale"].idxmax(), " - ", "${:,.2f}".format((produce_df.T)["Total Sale"].max()))
print()

## display orange and beets together for quantity sold and total sale using loc
print(produce_df.loc[['Quantity Sold','Total Sale'], ['Orange', 'Beets']])
print()

## us at to display total sales for cucumber
print("Total Cucumber Sales:","${:,.2f}".format(produce_df.at["Total Sale", "Cucumber"]))
print()

## New dataframe for only those produce that have sold between 11,500 to 12,000
t_df = produce_df.T
special_df = t_df[(t_df['Quantity Sold'] > 11499) & (t_df['Quantity Sold'] < 12001)]
special_df = special_df.T
print(special_df)
print()

## Make a transposed copy
t_special_df = special_df.T 
print(t_special_df)


#output to the screen
'''for key in produce_dict:
    
    print(key,"\t\t",
          '$',format(produce_dict[key]['cost'],',.2f'),"\t",
          format(produce_dict[key]['amt_sold'],',.2f'),"\t",
          '$',format(produce_dict[key]['total'],',.2f'))
          '''
##    
##
###write it to a .csv file instead
##
##    outfile = open("ProduceReport.csv", 'w')
##
##    outfile.write("Produce,Cost Per Pound,Amt Sold,Total\n")
##

##for key in produce_dict:
##        outfile.write(key + ',' +
##                      str(produce_dict[key]['cost']) + ',' +
##                      format(produce_dict[key]['amt_sold'],',.2f') + ',' +
##                      format(produce_dict[key]['total'],',.2f') + '\n'
##                      )

# why would we not want to do the above?
# becos format with ',' will create extra commas and cause problems reading
# the file

##for key in produce_dict:
##        outfile.write(key + ',' +
##                      str(produce_dict[key]['cost']) + ',' +
##                      str(round(produce_dict[key]['amt_sold'],2)) + ',' +
##                      str(round(produce_dict[key]['total'],2)) + '\n'
##                      )
##       
##        
##outfile.close()

    


    

    

    
    
